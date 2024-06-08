import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

file_path = 'data/第一次試驗/【測試結果】_60題.xlsx'
xls = pd.ExcelFile(file_path)

wb = openpyxl.load_workbook(file_path)

sheet_names = [f'alpha {i/10}' for i in range(10, 0, -1)]
for sheet in sheet_names:
    df = pd.read_excel(xls, sheet_name=sheet)
    
    E_col_group1 = pd.concat([df.iloc[:15, 4], df.iloc[30:45, 4]])
    E_col_group2 = pd.concat([df.iloc[15:30, 4], df.iloc[45:, 4]])
    F_col_group1 = pd.concat([df.iloc[:15, 5], df.iloc[30:45, 5]])
    F_col_group2 = pd.concat([df.iloc[15:30, 5], df.iloc[45:, 5]])

    # 語意 30 題
    E_yes_group1 = E_col_group1.value_counts().get('yes', 0)
    E_no_group1 = E_col_group1.value_counts().get('no', 0)
    F_yes_group1 = F_col_group1.value_counts().get('yes', 0)
    F_no_group1 = F_col_group1.value_counts().get('no', 0)

    if (E_yes_group1 + E_no_group1) > 0:
        E_percentage_group1 = E_yes_group1 / (E_yes_group1 + E_no_group1) * 100
    else:
        E_percentage_group1 = 0

    if (F_yes_group1 + F_no_group1) > 0:
        F_percentage_group1 = F_yes_group1 / (F_yes_group1 + F_no_group1) * 100
    else:
        F_percentage_group1 = 0

    # 關鍵字 30 題
    E_yes_group2 = E_col_group2.value_counts().get('yes', 0)
    E_no_group2 = E_col_group2.value_counts().get('no', 0)
    F_yes_group2 = F_col_group2.value_counts().get('yes', 0)
    F_no_group2 = F_col_group2.value_counts().get('no', 0)

    if (E_yes_group2 + E_no_group2) > 0:
        E_percentage_group2 = E_yes_group2 / (E_yes_group2 + E_no_group2) * 100
    else:
        E_percentage_group2 = 0

    if (F_yes_group2 + F_no_group2) > 0:
        F_percentage_group2 = F_yes_group2 / (F_yes_group2 + F_no_group2) * 100
    else:
        F_percentage_group2 = 0

    if sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb.create_sheet(title=sheet)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
    
    if ws.merged_cells:
        ws.unmerge_cells(str(ws.merged_cells.ranges))

    ws.cell(row=2, column=10).value = E_percentage_group1
    ws.cell(row=2, column=11).value = F_percentage_group1
    ws.cell(row=2, column=12).value = E_percentage_group2
    ws.cell(row=2, column=13).value = F_percentage_group2

wb.save('data/第一次試驗/【測試結果】_60題_含準確率.xlsx')
