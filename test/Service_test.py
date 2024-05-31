from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import requests, json, io, os
from openpyxl import load_workbook

from utils.weaviate_op import search_do
from utils.call_ai import call_aied
from utils.gemini_tem import Gemini_Template

app = Flask(__name__)
CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type, Qs-PageCode, Cache-Control'

url_template = "http://127.0.0.1:5000/chat?mess={question}&alpha={alpha}"

@app.route("/")
def index():
    """Server 是否正常的確認頁面."""
    return "server is ready"

@app.route("/chat", methods=['POST', 'GET'])
def chat_bot():
    question = request.values.get("mess")
    alpha = request.values.get("alpha")

    if not question:
        response = "無內容"
    else:
        try:
            response_li = search_do(question, alp=alpha)
            response = call_aied(response_li, question)
        except Exception as e:
            print(f"get error: {e}")
            response = f"Error: {e}"

    return jsonify({"llm": response, "retriv": response_li})

def verify_retriv_response(retriv_response, answer):
    for i in range(len(retriv_response)):
        retriv_response[i] = retriv_response[i].replace('"','').replace('\n','').replace('\\n','').replace('\\','').replace(' ','')
    return "yes" if answer in retriv_response else "no"

def verify_llm_response(llm_response, answer):
    prompt = f'問題: {llm_response} 與 {answer} 是否意思相同？是的話請回傳 yes，不是的話請回傳 no，只有這兩種回應，不要回其他東西'
    response = Gemini_Template(prompt)
    return "yes" if "yes" in response else "no"

@app.route('/llmautotest', methods=['POST'])
def process_excel():
    file = request.files['file']
    print("Received file:", file.filename)
    df = pd.read_excel(file)
    print("DataFrame loaded with columns:", df.columns.tolist())

    questions = df['問題'].tolist()
    answers = df['答案'].tolist()
    print("Extracted questions and answers")

    local_path = 'tmp/processed_results.xlsx'
    
    if os.path.exists(local_path):
        os.remove(local_path)
        print(f"Existing file at {local_path} removed")

    with pd.ExcelWriter(local_path, engine='openpyxl') as writer:
        df_placeholder = pd.DataFrame({'Placeholder': [1]})
        df_placeholder.to_excel(writer, sheet_name='placeholder', index=False)
    print(f"Initial file created with placeholder sheet at {local_path}")

    for alpha in [round(x * 0.1, 1) for x in range(10, 0, -1)]:
        print(f"Processing alpha: {alpha}")
        sheet_name = f'alpha {alpha}'
        for question, answer in zip(questions, answers):
            url = url_template.format(question=question, alpha=alpha)
            print(f"Generated URL: {url}")
            response = requests.get(url)
            if response.status_code == 200:
                result = response.json()
                retriv_result = result["retriv"]
                llm_result = result["llm"]
                retriv_validation = verify_retriv_response(retriv_result, answer.replace('"','').replace('\n','').replace('\\n','').replace('\\','').replace(' ',''))
                llm_validation = verify_llm_response(llm_result, answer)
                print(f"Results for question: {question}, alpha: {alpha} retrieved successfully")
            else:
                retriv_result = "Failed to get response"
                llm_result = "Failed to get response"
                retriv_validation = "no"
                llm_validation = "no"
                print(f"Failed to get response for question: {question} with alpha: {alpha}")

            result_df = pd.DataFrame({
                '問題': [question],
                '答案': [answer],
                f'檢索結果_{alpha}': [retriv_result],
                f'GPT_結果_{alpha}': [llm_result],
                f'檢索驗證_{alpha}': [retriv_validation],
                f'答案驗證_{alpha}': [llm_validation]
            })
            print(f"DataFrame created for question: {question}, alpha: {alpha}")

            with pd.ExcelWriter(local_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if sheet_name in writer.book.sheetnames:
                    startrow = writer.book[sheet_name].max_row
                    print(f"Appending to existing sheet: {sheet_name} starting at row {startrow}")
                else:
                    startrow = 0
                    print(f"Creating new sheet: {sheet_name}")

                result_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=(startrow == 0))
                print(f"DataFrame written to sheet: {sheet_name}")

    print(f"File saved to {local_path}")
    return send_file(local_path, download_name='processed_results.xlsx', as_attachment=True)

if __name__ == "__main__":
    if not os.path.exists('tmp'):
        os.makedirs('tmp')
    app.run(host="0.0.0.0", port=5000, threaded=True)
